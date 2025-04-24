import os
import pandas as pd

from openpyxl import load_workbook


def save_df2excel(df: pd.DataFrame, output_file: str = "output.xlsx", sheet_name: str = "data", index: bool = True) -> None:
    # ファイルの存在確認
    if not os.path.exists(output_file):
        # ファイルが存在しない場合、新規作成
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=index)
        print(f"新しいエクセルファイル:{output_file}, シート:{sheet_name}を作成しました。")
    else:
        # ファイルが存在する場合、追記処理
        try:
            book = load_workbook(output_file)

            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                writer.workbook = book # type: ignore
                writer.sheets.update({ws.title: ws for ws in book.worksheets})
                
                if sheet_name in book.sheetnames:
                    df.to_excel(writer, sheet_name=sheet_name, index=index, header=False, startrow=book[sheet_name].max_row)
                    print(f"エクセルファイル:{output_file}, シート:{sheet_name}にデータを追記しました。")
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=index)
                    print(f"エクセルファイル:{output_file}に、シート:{sheet_name}を作成しました。")
        except Exception as e:
            print(f"error: {e}")
